"""Заполнение шаблона чек-листа и акта по инвентарному номеру."""

from __future__ import annotations

import re
import shutil
import zipfile
from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd

from src.loaders import SourceData, load_all

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE = ROOT / "templates" / "чек-лист_акт_шаблон.xlsx"
DEFAULT_PEOPLE_FILE = ROOT / "templates" / ".people.xlsx"
DEFAULT_OUTPUT_DIR = ROOT / "output"

INV_COL = "Инвентарный номер АО УльГЭС"
TP_PATTERN = re.compile(r"(ТП|РП)-\d+", re.IGNORECASE)

# Лист «акт»: строки объектов 44–53 (10 шт.)
ACT_OBJECT_START = 44
ACT_OBJECT_END = 53
ACT_OBJECT_LABEL = "Кабельная линия"
ACT_UNIT = "км"
SPECIAL_OPINION_CELL = "B57"

# Лист «чек-лист»: строка 9 — объект; строки 12–20 — итог и блок представителей
CHECK_OBJECT_TEMPLATE_ROW = 9
CHECK_TAIL_START = 12
CHECK_TAIL_END = 20
CHECK_PRINT_AREA_END_COL = "R"
CHECK_LENGTH_SUM_COL = "D"  # ячейка с =SUM(N…)/1000 (исходно D12)
CHECK_LENGTH_DATA_COL = "N"  # колонка длин объектов
MAX_OBJECTS_PER_FILE = ACT_OBJECT_END - ACT_OBJECT_START + 1  # 10


@dataclass
class _TailBlockSnapshot:
    """Снимок строк 12–20 (итог + представители) до вставки объектов."""

    cells: dict[tuple[int, int], object] = field(default_factory=dict)  # (rel_row, col)
    styles: dict[tuple[int, int], tuple[object, object]] = field(
        default_factory=dict
    )  # (rel_row, col) -> (_style, number_format)
    row_heights: dict[int, object] = field(default_factory=dict)  # rel_row -> height
    merges: list[tuple[int, int, int, int]] = field(default_factory=list)  # rel rows


class ChecklistError(Exception):
    """Ошибка генерации чек-листа."""


@dataclass(frozen=True)
class ObjectRecord:
    """Один объект (строка списка кабелей) для инв. №."""

    инв_номер: int
    наименование: str
    адрес: str
    марка: str
    сечение: str
    длина_м: float
    год_ввода: int | None
    широта: float | None
    долгота: float | None
    тп: str | None


def _schedule_column(schedule: pd.DataFrame, *needles: str) -> str:
    """Ищет столбец графика по подстрокам в заголовке (регистронезависимо)."""
    for col in schedule.columns:
        if not isinstance(col, str):
            continue
        lower = col.lower()
        if all(n in lower for n in needles):
            return col
    raise ChecklistError(f"В графике не найден столбец ({', '.join(needles)})")


def abbreviate_fio(full_name: str) -> str:
    """Фамилия И. О. — как формула в ячейке B64 шаблона (сейчас не вызывается)."""
    parts = full_name.split()
    if not parts:
        return full_name
    surname = parts[0]
    initials = "".join(f" {p[0]}." for p in parts[1:3] if p)
    return f"{surname}{initials}"


def extract_tp_name(text: str) -> str | None:
    """Извлекает «ТП-…» или «РП-…» из диспетчерского наименования."""
    match = TP_PATTERN.search(text)
    if not match:
        return None
    return match.group(0).upper().replace("тп", "ТП").replace("рп", "РП")


def _lookup_substation(substations: pd.DataFrame, tp_name: str) -> tuple[float | None, float | None]:
    """Координаты подстанции: сначала точное совпадение Name, затем вхождение подстроки."""
    names = substations["Name"].astype(str)
    row = substations[names.str.upper() == tp_name.upper()]
    if row.empty:
        row = substations[names.str.contains(tp_name, case=False, na=False)]
    if row.empty:
        return None, None
    rec = row.iloc[0]
    return rec.get("широта"), rec.get("долгота")


def _cable_row_to_object(
    row: pd.Series,
    inv: int,
    schedule_name: str,
    адрес: str,
    substations: pd.DataFrame,
) -> ObjectRecord:
    """Собирает ObjectRecord из строки списка кабелей и общих полей графика."""
    # Наименование из списка кабелей приоритетнее диспетчерского из графика.
    наименование = schedule_name
    if "наименование" in row.index and pd.notna(row["наименование"]):
        candidate = str(row["наименование"]).strip()
        if candidate:
            наименование = candidate

    длина = pd.to_numeric(row["длина"], errors="coerce")
    год = row.get("год_ввода")
    год_ввода = int(год) if pd.notna(год) else None

    тп = extract_tp_name(наименование)
    широта, долгота = (None, None)
    if тп:
        широта, долгота = _lookup_substation(substations, тп)

    return ObjectRecord(
        инв_номер=inv,
        наименование=наименование,
        адрес=адрес,
        марка=str(row["марка"]),
        сечение=str(row["сечение"]),
        длина_м=float(длина),
        год_ввода=год_ввода,
        широта=широта,
        долгота=долгота,
        тп=тп,
    )


def collect_objects(data: SourceData, inv: int) -> list[ObjectRecord]:
    """Все объекты (строки кабелей) для инв. №."""
    schedule = data.schedule
    if INV_COL not in schedule.columns:
        raise ChecklistError(f"В графике нет столбца {INV_COL!r}")

    # Шапка акта/чек-листа — одна строка графика на инв. №.
    sch_rows = schedule[schedule[INV_COL] == inv]
    if sch_rows.empty:
        raise ChecklistError(f"Инв. № {inv} не найден в графике инвентаризации")

    sch_row = sch_rows.iloc[0]
    name_col = _schedule_column(schedule, "диспетчер")
    address_col = _schedule_column(schedule, "адрес", "местонахожд")
    schedule_name = str(sch_row[name_col]).strip()
    адрес = str(sch_row[address_col]).strip()

    cab_rows = data.cables[data.cables["инв_номер"] == inv]
    if cab_rows.empty:
        raise ChecklistError(f"Инв. № {inv} не найден в списке кабелей")

    # Каждая строка кабеля — отдельный объект в таблицах акта и чек-листа.
    objects = [
        _cable_row_to_object(row, inv, schedule_name, адрес, data.substations)
        for _, row in cab_rows.iterrows()
    ]

    missing_coords = {obj.тп for obj in objects if obj.тп and obj.широта is None}
    for тп in missing_coords:
        print(f"Предупреждение: координаты для {тп} не найдены в справочнике подстанций")

    return objects


def _set(ws, addr: str, value) -> None:
    """Запись значения в ячейку по адресу (A1-нотация)."""
    ws[addr] = value


def _clear_merges_in_rows(ws, row_from: int, row_to: int) -> None:
    """Снимает объединения ячеек, пересекающиеся с диапазоном строк."""
    for merge in list(ws.merged_cells.ranges):
        if merge.min_row <= row_to and merge.max_row >= row_from:
            # На пустых (вставленных) строках ячеек объединения может не быть —
            # создаём их, иначе unmerge_cells падает с KeyError.
            for row in range(merge.min_row, merge.max_row + 1):
                for col in range(merge.min_col, merge.max_col + 1):
                    ws.cell(row, col)
            ws.unmerge_cells(str(merge))


def _clear_sheet_selection(ws) -> None:
    """Убирает выделение строки/диапазона при открытии файла."""
    if ws.sheet_view is not None:
        ws.sheet_view.selection = []


def _extend_check_print_area(ws, extra_rows: int) -> None:
    """Сдвигает нижнюю границу области печати после вставки строк объектов."""
    if extra_rows <= 0:
        return
    ws.print_area = f"A1:{CHECK_PRINT_AREA_END_COL}{CHECK_TAIL_END + extra_rows}"


def _update_length_sum_formula(ws, count: int, offset: int) -> None:
    """Обновляет формулу суммы длин (исходно D12) под фактические строки объектов."""
    first = CHECK_OBJECT_TEMPLATE_ROW
    last = CHECK_OBJECT_TEMPLATE_ROW + count - 1
    sum_row = CHECK_TAIL_START + offset
    ws[f"{CHECK_LENGTH_SUM_COL}{sum_row}"] = (
        f"=SUM({CHECK_LENGTH_DATA_COL}{first}:{CHECK_LENGTH_DATA_COL}{last})/1000"
    )


def _snapshot_tail_block(ws) -> _TailBlockSnapshot:
    """Сохраняет ячейки и объединения строк 12–20 (итог + представители)."""
    snap = _TailBlockSnapshot()
    base = CHECK_TAIL_START
    for merge in ws.merged_cells.ranges:
        if merge.min_row >= base and merge.max_row <= CHECK_TAIL_END:
            snap.merges.append(
                (merge.min_row - base, merge.min_col, merge.max_row - base, merge.max_col)
            )
    for row in range(base, CHECK_TAIL_END + 1):
        snap.row_heights[row - base] = ws.row_dimensions[row].height
        for col in range(1, 18):  # до колонки Q
            cell = ws.cell(row, col)
            snap.styles[(row - base, col)] = (copy(cell._style), cell.number_format)
            if cell.__class__.__name__ != "MergedCell" and cell.value is not None:
                snap.cells[(row - base, col)] = cell.value
    return snap


def _restore_tail_block(ws, snap: _TailBlockSnapshot, offset: int) -> None:
    """Восстанавливает блок представителей на строках 12+offset … 20+offset."""
    if offset <= 0:
        return

    # Устаревшие merge зоны 12–20 уже сняты ранее; здесь чистим только приёмник,
    # иначе при count > 3 затрём объединения строк-объектов, попавших в 12–20.
    dst_start = CHECK_TAIL_START + offset
    dst_end = CHECK_TAIL_END + offset
    _clear_merges_in_rows(ws, dst_start, dst_end)

    for (rel_row, col), (style, number_format) in snap.styles.items():
        dst = ws.cell(CHECK_TAIL_START + rel_row + offset, col)
        dst._style = copy(style)
        dst.number_format = number_format

    for rel_row, height in snap.row_heights.items():
        ws.row_dimensions[CHECK_TAIL_START + rel_row + offset].height = height

    for (rel_row, col), value in snap.cells.items():
        ws.cell(CHECK_TAIL_START + rel_row + offset, col).value = value

    for r1, c1, r2, c2 in snap.merges:
        ws.merge_cells(
            start_row=CHECK_TAIL_START + r1 + offset,
            end_row=CHECK_TAIL_START + r2 + offset,
            start_column=c1,
            end_column=c2,
        )


def _copy_row_style(ws, template_row: int, target_row: int, min_col: int, max_col: int) -> None:
    """Копирует высоту строки и стиль ячеек из образца (без значений)."""
    ws.row_dimensions[target_row].height = ws.row_dimensions[template_row].height
    for col in range(min_col, max_col + 1):
        src = ws.cell(template_row, col)
        dst = ws.cell(target_row, col)
        dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format


def _copy_row_merges(ws, template_row: int, target_row: int) -> None:
    """Дублирует горизонтальные объединения ячеек строки-образца."""
    for merge in list(ws.merged_cells.ranges):
        if merge.min_row == template_row == merge.max_row:
            ws.merge_cells(
                start_row=target_row,
                end_row=target_row,
                start_column=merge.min_col,
                end_column=merge.max_col,
            )


def _copy_check_row_formulas(ws, template_row: int, target_row: int) -> None:
    """Копирует формулы L и M из строки-образца (например ``=D5``)."""
    for col in (12, 13):  # L, M
        src = ws.cell(template_row, col)
        value = src.value
        if isinstance(value, str) and value.startswith("="):
            ws.cell(target_row, col).value = value


def _ensure_check_object_rows(ws, count: int) -> list[int]:
    """
    Лист осмотра: строка 9 — образец; при count > 1 вставляет строки после 9-й.
    Возвращает номера строк объектов (не более 10).
    """
    if count < 1 or count > MAX_OBJECTS_PER_FILE:
        raise ChecklistError(
            f"Число объектов на лист должно быть 1–{MAX_OBJECTS_PER_FILE}, получено {count}"
        )

    template = CHECK_OBJECT_TEMPLATE_ROW
    if count > 1:
        tail_snap = _snapshot_tail_block(ws)
        insert_at = template + 1
        _clear_merges_in_rows(ws, insert_at, insert_at)
        ws.insert_rows(insert_at, count - 1)
        offset = count - 1
        # insert_rows не сдвигает объединения: чистим устаревшие merge до копирования
        # стиля, иначе unmerge_cells сбросит форматирование вставленных строк.
        _clear_merges_in_rows(ws, insert_at, CHECK_TAIL_END)
        for row in range(template + 1, template + count):
            _copy_row_style(ws, template, row, 1, 22)
            _copy_row_merges(ws, template, row)
            _copy_check_row_formulas(ws, template, row)
        _restore_tail_block(ws, tail_snap, offset)
        _extend_check_print_area(ws, offset)

    _update_length_sum_formula(ws, count, count - 1)
    _clear_sheet_selection(ws)
    return list(range(template, template + count))


def _prepare_act_object_row(ws, act_row: int) -> None:
    """Готовит строку объекта на листе акта: стиль и merge E:F (кроме первой строки 44)."""
    if act_row == ACT_OBJECT_START:
        return
    if act_row > ACT_OBJECT_END:
        raise ChecklistError(
            f"На листе акта доступны строки {ACT_OBJECT_START}–{ACT_OBJECT_END}, "
            f"требуется строка {act_row}"
        )
    _copy_row_style(ws, ACT_OBJECT_START, act_row, 2, 9)
    ws.merge_cells(start_row=act_row, end_row=act_row, start_column=5, end_column=6)


def _special_opinion_text(
    *,
    inv: int,
    part: int,
    total_parts: int,
    obj_from: int,
    obj_to: int,
    total_objects: int,
    act_number: int | str | None,
) -> str:
    """Текст в B57 при разбиении инв. № на несколько файлов (продолжение акта)."""
    act_part = f" № {act_number}" if act_number is not None else ""
    if part == 1:
        return (
            f"Инв. № {inv}. Акт{act_part}. Объекты {obj_from}–{obj_to} из {total_objects}. "
            f"Продолжение — файл чек-лист_{inv}_{part + 1}.xlsx."
        )
    if part == total_parts:
        return (
            f"Инв. № {inv}. Акт{act_part} (продолжение). "
            f"Объекты {obj_from}–{obj_to} из {total_objects}."
        )
    return (
        f"Инв. № {inv}. Акт{act_part} (продолжение). Объекты {obj_from}–{obj_to} из {total_objects}. "
        f"Продолжение — файл чек-лист_{inv}_{part + 1}.xlsx."
    )


def copy_people_to_output(
    output_dir: Path,
    *,
    source: Path | None = None,
) -> Path:
    """Копирует ``templates/.people.xlsx`` в каталог вывода (внешняя ссылка шаблона)."""
    src = source or DEFAULT_PEOPLE_FILE
    if not src.is_file():
        raise ChecklistError(f"Справочник представителей не найден: {src}")

    output_dir.mkdir(parents=True, exist_ok=True)
    dest = output_dir / src.name
    if not dest.is_file() or src.stat().st_mtime > dest.stat().st_mtime:
        shutil.copy2(src, dest)
    return dest


def _output_path(out_dir: Path, inv: int, part: int, total_parts: int) -> Path:
    """Имя выходного файла: без суффикса при одном файле, иначе ``_N``."""
    if total_parts == 1:
        return out_dir / f"чек-лист_{inv}.xlsx"
    return out_dir / f"чек-лист_{inv}_{part}.xlsx"


def _with_copy_suffix(path: Path, copy_index: int) -> Path:
    """Имя файла-копии, как в проводнике Windows: «… - Копия», «… - Копия (2)»."""
    if copy_index == 0:
        return path
    stem = path.stem
    suffix = path.suffix
    if copy_index == 1:
        return path.with_name(f"{stem} - Копия{suffix}")
    return path.with_name(f"{stem} - Копия ({copy_index}){suffix}")


def _merge_content_types(template_xml: str, output_xml: str) -> str:
    """Добавляет в ``[Content_Types].xml`` записи из шаблона, которых нет в output."""
    merged = output_xml
    for pattern in (
        r'<Default Extension="bin"[^>]*/>',
        r'<Override PartName="/xl/drawings/[^"]+"[^>]*/>',
    ):
        for match in re.finditer(pattern, template_xml):
            tag = match.group(0)
            if tag not in merged:
                merged = merged.replace("</Types>", f"{tag}</Types>")
    return merged


_RELATIONSHIPS_NS = (
    'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
)


def _ensure_relationships_ns(worksheet_xml: str) -> str:
    """openpyxl не объявляет ``xmlns:r``, но ``pageSetup``/``drawing`` ссылаются через ``r:id``."""
    if "xmlns:r=" in worksheet_xml or "r:id" not in worksheet_xml:
        return worksheet_xml
    return re.sub(
        r"(<worksheet\b[^>]*)(>)",
        lambda m: (
            m.group(1)
            + ("" if "xmlns:r=" in m.group(1) else f" {_RELATIONSHIPS_NS}")
            + m.group(2)
        ),
        worksheet_xml,
        count=1,
    )


def _patch_worksheet_xml(template_xml: str, output_xml: str, *, sheet: str) -> str:
    """Восстанавливает ``pageSetup r:id`` и ``drawing`` из шаблона."""
    patched = output_xml
    page_setup = re.search(r"<pageSetup[^>]*/>", template_xml)
    if page_setup:
        patched = re.sub(r"<pageSetup[^>]*/>", page_setup.group(0), patched, count=1)

    if sheet == "sheet1":
        drawing = re.search(r'<drawing r:id="[^"]+"\s*/>', template_xml)
        if drawing and "drawing" not in patched:
            patched = patched.replace("</worksheet>", f"{drawing.group(0)}</worksheet>")

    return _ensure_relationships_ns(patched)


def _restore_template_package(template_path: Path, output_path: Path) -> None:
    """Возвращает в output части xlsx из шаблона, которые openpyxl не сохраняет."""
    preserve_prefixes = ("xl/drawings/", "xl/printerSettings/")
    preserve_files = (
        "xl/worksheets/_rels/sheet1.xml.rels",
        "xl/worksheets/_rels/sheet2.xml.rels",
    )

    with zipfile.ZipFile(template_path) as ztpl, zipfile.ZipFile(output_path) as zout:
        tpl = {info.filename: ztpl.read(info.filename) for info in ztpl.infolist()}
        out_names = [info.filename for info in zout.infolist()]
        out = {name: zout.read(name) for name in out_names}

    for name, data in tpl.items():
        if name.startswith(preserve_prefixes) or name in preserve_files:
            out[name] = data

    for sheet in ("sheet1", "sheet2"):
        sheet_name = f"xl/worksheets/{sheet}.xml"
        out[sheet_name] = _patch_worksheet_xml(
            tpl[sheet_name].decode("utf-8"),
            out[sheet_name].decode("utf-8"),
            sheet=sheet,
        ).encode("utf-8")

    out["[Content_Types].xml"] = _merge_content_types(
        tpl["[Content_Types].xml"].decode("utf-8"),
        out["[Content_Types].xml"].decode("utf-8"),
    ).encode("utf-8")

    with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in out.items():
            zout.writestr(name, data)


def _write_checklist_file(
    template_path: Path,
    desired_path: Path,
    *,
    fill_fn,
) -> Path:
    """Пишет чек-лист; при занятом файле — в «… - Копия», как в проводнике Windows."""
    copy_index = 0
    while copy_index < 100:
        output_path = _with_copy_suffix(desired_path, copy_index)
        try:
            if copy_index > 0:
                print(
                    f"Файл занят: {desired_path.name} — сохраняем как {output_path.name}",
                )
            shutil.copy2(template_path, output_path)
            wb = openpyxl.load_workbook(output_path)
            fill_fn(wb)
            wb.save(output_path)
            _restore_template_package(template_path, output_path)
            return output_path
        except PermissionError:
            copy_index += 1
    raise ChecklistError(f"Не удалось сохранить файл (занят): {desired_path.name}")


def fill_act_object_row(ws, act_row: int, seq: int, obj: ObjectRecord) -> None:
    """Заполняет одну строку таблицы объектов на листе «акт» (строки 44–53)."""
    _prepare_act_object_row(ws, act_row)
    _set(ws, f"B{act_row}", seq)
    _set(ws, f"C{act_row}", ACT_OBJECT_LABEL)
    _set(ws, f"D{act_row}", obj.марка)
    _set(ws, f"E{act_row}", obj.сечение)
    _set(ws, f"G{act_row}", ACT_UNIT)
    _set(ws, f"H{act_row}", obj.длина_м / 1000)  # шаблон: длина в км
    _set(ws, f"I{act_row}", obj.адрес)


def fill_check_object_row(ws, chk_row: int, seq: int, obj: ObjectRecord) -> None:
    """Заполняет одну строку объекта на листе «чек-лист» (колонки A–N)."""
    _set(ws, f"A{chk_row}", seq)
    _set(ws, f"B{chk_row}", obj.наименование)
    _set(ws, f"E{chk_row}", obj.инв_номер)
    _set(ws, f"J{chk_row}", obj.марка)
    _set(ws, f"K{chk_row}", obj.сечение)
    _set(ws, f"N{chk_row}", obj.длина_м)

    if obj.год_ввода is not None:
        _set(ws, f"D{chk_row}", obj.год_ввода)
    if obj.широта is not None:
        _set(ws, f"F{chk_row}", obj.широта)
    if obj.долгота is not None:
        _set(ws, f"G{chk_row}", obj.долгота)


def fill_workbook(
    wb: openpyxl.Workbook,
    objects: list[ObjectRecord],
    *,
    act_number: int | str | None = None,
    act_date: date | datetime | None = None,
    part: int = 1,
    total_parts: int = 1,
    global_offset: int = 0,
    total_objects: int | None = None,
) -> None:
    """
    Заполняет оба листа шаблона: акт (лист 0) и чек-лист (лист 1).

    ``global_offset`` — сквозной номер первого объекта при многофайловой генерации.
    """
    act = wb.worksheets[0]  # лист «акт»
    chk = wb.worksheets[1]  # лист «чек-лист»
    total_objects = total_objects or len(objects)
    count = len(objects)

    # Реквизиты акта
    if act_number is not None:
        _set(act, "E3", act_number)
    if act_date is not None:
        _set(act, "G3", act_date)

    # Общие поля шапки (одинаковы для всех объектов инв. №)
    header = objects[0]
    _set(act, "D7", header.наименование)
    _set(act, "D16", header.адрес)

    # Сокращённое ФИО в B64 отключено — остаётся значение/формула шаблона.
    # h20 = act["H20"].value
    # if h20:
    #     _set(act, "B64", abbreviate_fio(str(h20).strip()))

    _set(chk, "I4", header.наименование)
    _set(chk, "G4", header.адрес)

    # При count > 1 вставляет строки объектов и сдвигает блок представителей вниз.
    check_rows = _ensure_check_object_rows(chk, count)

    # Таблицы объектов на обоих листах
    for i, obj in enumerate(objects):
        seq = global_offset + i + 1
        act_row = ACT_OBJECT_START + i
        fill_act_object_row(act, act_row, seq, obj)
        fill_check_object_row(chk, check_rows[i], seq, obj)

    if total_parts > 1:
        # Пометка о продолжении акта в соседнем файле
        obj_from = global_offset + 1
        obj_to = global_offset + count
        _set(
            act,
            SPECIAL_OPINION_CELL,
            _special_opinion_text(
                inv=header.инв_номер,
                part=part,
                total_parts=total_parts,
                obj_from=obj_from,
                obj_to=obj_to,
                total_objects=total_objects,
                act_number=act_number,
            ),
        )


def generate_checklist(
    inv: int,
    *,
    output_dir: Path | None = None,
    template: Path | None = None,
    act_number: int | str | None = None,
    act_date: date | datetime | None = None,
    data: SourceData | None = None,
) -> list[Path]:
    """Создаёт один или несколько файлов ``чек-лист_{инв}[_N].xlsx``."""
    template_path = template or DEFAULT_TEMPLATE
    if not template_path.is_file():
        raise ChecklistError(f"Шаблон не найден: {template_path}")

    out_dir = output_dir or DEFAULT_OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    copy_people_to_output(out_dir)

    source_data = data or load_all()
    all_objects = collect_objects(source_data, inv)
    total_objects = len(all_objects)

    # Больше MAX_OBJECTS_PER_FILE — несколько файлов с продолжением акта
    if total_objects > MAX_OBJECTS_PER_FILE:
        total_parts = (total_objects + MAX_OBJECTS_PER_FILE - 1) // MAX_OBJECTS_PER_FILE
        print(
            f"Инв. № {inv}: найдено {total_objects} объектов — "
            f"будет создано {total_parts} файлов (до {MAX_OBJECTS_PER_FILE} объектов в каждом)"
        )
    else:
        total_parts = 1

    created: list[Path] = []

    # Каждая часть — копия шаблона и свой срез объектов
    for part in range(1, total_parts + 1):
        start = (part - 1) * MAX_OBJECTS_PER_FILE
        chunk = all_objects[start : start + MAX_OBJECTS_PER_FILE]
        desired_path = _output_path(out_dir, inv, part, total_parts)
        output_path = _write_checklist_file(
            template_path,
            desired_path,
            fill_fn=lambda wb: fill_workbook(
                wb,
                chunk,
                act_number=act_number,
                act_date=act_date,
                part=part,
                total_parts=total_parts,
                global_offset=start,
                total_objects=total_objects,
            ),
        )
        created.append(output_path)

    return created
